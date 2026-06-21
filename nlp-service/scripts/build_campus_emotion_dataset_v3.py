import csv
import io
from collections import Counter
from zipfile import is_zipfile
from pathlib import Path

from openpyxl import load_workbook


LABELS = [
    "lonely",
    "anxious",
    "stressed",
    "tired",
    "sad",
    "calm",
    "healed",
    "secure",
    "happy",
    "hopeful",
]

SOURCE_FILES = [
    "campus_emotion_samples_v2.csv",
    "manual_samples_王宇航.csv",
    "manual_samples_马振杰.csv",
    "manual_samples_郭柯嘉.csv",
    "manual_samples_张宇翔.csv",
    "manual_samples_艾科山·斯拉木.csv",
    "manual_samples_丁丹丹.csv",
]

ENCODINGS = ["utf-8-sig", "utf-8", "gb18030", "gbk"]
REQUIRED_FIELDS = ["id", "text", "emotion_type", "source"]

MA_ZHENJIE_TEXT_FIXES = {
    "1": "身边同学都开始实习了，我也开始担心自己是不是该去找实习",
    "7": "有时候觉得所有选择都可能后悔，但还是得自己做决定",
    "8": "顺其自然一点，也许现在经历的事都是在为以后做准备",
    "10": "普通的一天里突然有了不一样的小快乐，整个人都轻了一点",
    "13": "一个人去陌生的地方走走，也是一段很珍贵的经历",
    "14": "独自去参加活动时和陌生人打招呼，认识新朋友会让我很开心",
    "20": "又想吃点喜欢的东西安慰自己，又被各种计划和要求拉扯",
    "21": "可能我注定不能和喜欢的人在一起，想到这里心里很难过",
    "22": "原来那些不堪的起点，也可能在后来变成支撑自己的地方",
    "23": "等慢慢找到真正的自己，也许会重新觉得世界和生命都很美好",
    "24": "只要我不伤害自己，就想试着把自己保护好一点",
    "25": "好朋友也不一定要被捆绑着一起走，不被顾及的关系会让人很累",
    "27": "一个人吃饭其实还好，就是有时候突然不知道能找谁说话",
    "28": "宿舍里大家都有自己的圈子，我好像一直插不进去",
    "31": "刷实习经验刷到头疼，越看越焦虑",
    "33": "ddl 真的会把人逼疯，刚写完一个又来一个",
    "36": "有些关系慢慢淡掉的时候，还是会有点难过",
    "40": "今天状态一般，但也没有很糟，算平稳过完了",
    "41": "下雨后空气好舒服，走回宿舍那段路心情好了一点",
    "42": "喝到热奶茶的时候真的缓过来一点，今天太冷了",
    "44": "朋友听我碎碎念了半天也没嫌烦，真的会安心一点",
    "46": "本来很慌，结果老师说可以慢慢改，终于能喘口气",
    "48": "今天和朋友出去吃了顿饭，没干啥但就是挺开心",
    "49": "下班路上买了疯狂星期四，突然觉得今天也还不错",
    "50": "虽然现在挺乱的，但感觉慢慢来应该也不是完全没希望",
}


def decode_csv_text(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ENCODINGS:
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue

    # One submitted manual file was saved with a few invalid UTF-8 bytes and
    # shifted labels like "...?anxious,manual,manual". Read it lossily, then
    # repair only rows where the label is still unambiguous.
    if path.name == "manual_samples_马振杰.csv":
        return raw.decode("utf-8", errors="replace"), "utf-8-repaired"

    raise ValueError(f"cannot decode {path}")


def repair_row(path: Path, row: dict[str, str]) -> dict[str, str]:
    if path.name != "manual_samples_马振杰.csv":
        return row

    if row["emotion_type"] == "manual" and row["source"] == "manual":
        suffix_labels = [label for label in LABELS if row["text"].endswith(label)]
        if len(suffix_labels) == 1:
            label = suffix_labels[0]
            row["text"] = row["text"][: -len(label)].rstrip(" ,，?？�")
            row["emotion_type"] = label
            row["_repaired"] = "true"

    fixed_text = MA_ZHENJIE_TEXT_FIXES.get(row["id"])
    if fixed_text:
        row["text"] = fixed_text
        row["_repaired"] = "true"
    return row


def read_csv_any_encoding(path: Path) -> tuple[list[dict[str, str]], str]:
    if is_zipfile(path):
        return read_excel_content(path)

    text, encoding = decode_csv_text(path)
    reader = csv.DictReader(io.StringIO(text))
    fields = [(field or "").strip().lstrip("\ufeff") for field in (reader.fieldnames or [])]
    if fields != REQUIRED_FIELDS:
        raise ValueError(f"{path.name}: fields must be {REQUIRED_FIELDS}, got {fields}")

    rows: list[dict[str, str]] = []
    for row in reader:
        normalized = {
            (key or "").strip().lstrip("\ufeff"): (value or "").strip()
            for key, value in row.items()
        }
        if not any(normalized.values()):
            continue
        rows.append(repair_row(path, normalized))
    return rows, encoding


def read_excel_content(path: Path) -> tuple[list[dict[str, str]], str]:
    workbook = load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise ValueError(f"{path.name}: no rows")

    fields = [(str(field).strip() if field is not None else "") for field in raw_rows[0]]
    if fields != REQUIRED_FIELDS:
        raise ValueError(f"{path.name}: fields must be {REQUIRED_FIELDS}, got {fields}")

    rows: list[dict[str, str]] = []
    for raw_row in raw_rows[1:]:
        normalized = {
            field: ("" if value is None else str(value).strip())
            for field, value in zip(REQUIRED_FIELDS, raw_row)
        }
        if not any(normalized.values()):
            continue
        rows.append(repair_row(path, normalized))
    return rows, "xlsx-content"


def validate_rows(name: str, rows: list[dict[str, str]]) -> None:
    if not rows:
        raise ValueError(f"{name}: no rows")
    for index, row in enumerate(rows, start=2):
        if not row["text"]:
            raise ValueError(f"{name}: line {index} empty text")
        if row["emotion_type"] not in LABELS:
            raise ValueError(f"{name}: line {index} invalid emotion_type {row['emotion_type']!r}")
        if row["source"] not in {"manual", "ai_generated"}:
            raise ValueError(f"{name}: line {index} invalid source {row['source']!r}")


def write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=REQUIRED_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_excel_friendly_copy(path: Path, rows: list[dict[str, str]]) -> None:
    output = path.with_name(path.stem + "_excel_utf8_bom.csv")
    with output.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=REQUIRED_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx_copy(path: Path, rows: list[dict[str, str]]) -> None:
    output = path.with_suffix(".xlsx")
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = path.stem[:31]
    sheet.append(REQUIRED_FIELDS)
    for row in rows:
        sheet.append([row[field] for field in REQUIRED_FIELDS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = {"A": 8, "B": 72, "C": 16, "D": 16}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(horizontal="center")
        row[3].alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    data_dir = root / "data"

    merged: list[dict[str, str]] = []
    for file_name in SOURCE_FILES:
        path = data_dir / file_name
        rows, encoding = read_csv_any_encoding(path)
        validate_rows(file_name, rows)
        repaired_count = sum(1 for row in rows if row.get("_repaired") == "true")
        repair_note = f", repaired={repaired_count}" if repaired_count else ""
        print(
            f"{file_name}: {len(rows)} rows encoding={encoding} "
            f"{dict(Counter(row['source'] for row in rows))}{repair_note}"
        )
        merged.extend(rows)

    text_counts = Counter(row["text"] for row in merged)
    duplicate_texts = [text for text, count in text_counts.items() if count > 1]
    if duplicate_texts:
        raise ValueError(f"duplicate texts found: {duplicate_texts[:5]}")

    output_rows = [
        {
            "id": str(index),
            "text": row["text"],
            "emotion_type": row["emotion_type"],
            "source": row["source"],
        }
        for index, row in enumerate(merged, start=1)
    ]

    output = data_dir / "campus_emotion_samples_v3.csv"
    write_rows(output, output_rows)
    write_excel_friendly_copy(output, output_rows)
    write_xlsx_copy(output, output_rows)

    print(f"wrote {len(output_rows)} rows to {output}")
    print("emotion_counts:", dict(Counter(row["emotion_type"] for row in output_rows)))
    print("source_counts:", dict(Counter(row["source"] for row in output_rows)))


if __name__ == "__main__":
    main()
